#!/usr/bin/env python3
"""Uzupelnia brakujace powiat_ujednolicony i powiat_uwagi w jednym JSON-ie.

"""

from __future__ import annotations

import argparse
from collections import defaultdict
import csv
from datetime import datetime, timezone
import json
from pathlib import Path
import re
import sys
import time
from typing import Any, Iterator

import openpyxl


SOURCE_FIELD = "powiat_ocr"
TARGET_FIELD = "powiat_ujednolicony"
NOTE_FIELD = "powiat_uwagi"
NOTE_NOT_APPLICABLE = "nie dotyczy"

BASE_PATH = Path(__file__).resolve().parent.parent
DEFAULT_PRIMARY_XLSX = (
    BASE_PATH
    / "dane_do_korekt"
    / "powiaty_ujednolicon 13.01.2026 r. wersja uzupełniona - nazwa_json_org.xlsx"
)
DEFAULT_FALLBACK_XLSX = (
    BASE_PATH / "dane_do_korekt" / "powiaty_ujednolicone.xlsx"
)
DEFAULT_OUTPUT_DIR = BASE_PATH / "sgkp_uzupelnienie" / "powiaty"


def normalize_header(value: Any) -> str:
    """Normalizuje naglowek arkusza do porownan niezaleznych od spacji."""

    return re.sub(r"\s+", "_", str(value or "").strip().casefold())


def normalize_lookup_key(value: Any) -> str:
    """Normalizuje klucz tylko w bezpiecznym zakresie: spacje i casefold."""

    return " ".join(str(value or "").split()).casefold()


def clean_source_value(value: Any) -> str:
    return " ".join(str(value or "").split()).strip()


def has_value(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    return value.strip().casefold() not in {"", "/", "null", "none", "brak"}


def normalize_powiat(value: Any) -> str:
    """Usuwa z nazwy ujednoliconej doprecyzowania techniczne w nawiasach."""

    text = str(value or "")
    text = re.sub(r"\s*\([^)]*\)", "", text)
    text = text.replace("(", "").replace(")", "")
    text = re.sub(
        r"\s*[-–]?\s*\bnie dotyczy\b", "", text, flags=re.IGNORECASE
    )
    return " ".join(text.split()).strip()


def normalize_note(value: Any) -> str:
    if isinstance(value, str) and re.search(
        r"\bnie dotyczy\b", value, flags=re.IGNORECASE
    ):
        return NOTE_NOT_APPLICABLE
    return ""


def parse_mapping_outcome(unified_value: Any, note_value: Any) -> tuple[str, str]:
    """Laczy nazwe ujednolicona z informacja o zakresie projektu."""

    powiat = normalize_powiat(unified_value)
    note = normalize_note(note_value) or normalize_note(unified_value)
    return powiat, note


def read_headers(sheet: Any) -> dict[str, int]:
    first_row = next(sheet.iter_rows(min_row=1, max_row=1))
    return {
        normalize_header(cell.value): index
        for index, cell in enumerate(first_row)
        if cell.value is not None
    }


def read_mapping_rows(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Wczytuje niepuste wiersze mapowania z jednego arkusza XLSX."""

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers = read_headers(sheet)
        required = {"nazwa_json", "nazwa_ujednolicona"}
        missing = sorted(required - set(headers))
        if missing:
            raise ValueError(
                f"Brak kolumn w arkuszu {path}: {', '.join(missing)}"
            )
        relevant_columns = [
            headers[name]
            for name in (
                "nazwa_json",
                "nazwa_json_org",
                "nazwa_ujednolicona",
                "dotyczy?",
                "identyfikator",
            )
            if name in headers
        ]
        max_col = max(relevant_columns) + 1

        rows: list[dict[str, Any]] = []
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, max_col=max_col, values_only=True),
            start=2,
        ):
            def cell(column: str) -> Any:
                index = headers.get(column)
                return row[index] if index is not None and index < len(row) else None

            unified = cell("nazwa_ujednolicona")
            name_json = clean_source_value(cell("nazwa_json"))
            name_org = clean_source_value(cell("nazwa_json_org"))
            identifier = clean_source_value(cell("identyfikator"))
            if unified is None or not (name_json or name_org or identifier):
                continue
            outcome = parse_mapping_outcome(unified, cell("dotyczy?"))
            if not any(outcome):
                continue
            rows.append(
                {
                    "row_number": row_number,
                    "nazwa_json": name_json,
                    "nazwa_json_org": name_org,
                    "identyfikator": identifier,
                    "powiat": outcome[0],
                    "uwagi": outcome[1],
                }
            )
        metadata = {
            "plik": str(path),
            "arkusz": sheet.title,
            "wiersze_mapowania": len(rows),
            "kolumny": sorted(headers),
        }
        return rows, metadata
    finally:
        workbook.close()


def group_rows_by_key(
    rows: list[dict[str, Any]],
    key_column: str,
) -> dict[str, dict[tuple[str, str], list[int]]]:
    groups: dict[str, dict[tuple[str, str], list[int]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for row in rows:
        key = row.get(key_column, "")
        if not key:
            continue
        outcome = (row["powiat"], row["uwagi"])
        groups[key][outcome].append(int(row["row_number"]))
    return groups


def add_grouped_mappings(
    catalog: dict[str, dict[str, Any]],
    blocked_keys: set[str],
    conflicts: list[dict[str, Any]],
    groups: dict[str, dict[tuple[str, str], list[int]]],
    source_path: Path,
    key_column: str,
    source_role: str,
) -> None:
    """Dodaje tylko nowe, jednoznaczne klucze zgodnie z priorytetem."""

    for key, variants in groups.items():
        if key in catalog or key in blocked_keys:
            continue
        if len(variants) != 1:
            blocked_keys.add(key)
            conflicts.append(
                {
                    "rodzaj_konfliktu": "dokladny_klucz",
                    "klucz": key,
                    "kolumna": key_column,
                    "zrodlo": str(source_path),
                    "warianty": [
                        {
                            "powiat": outcome[0],
                            "uwagi": outcome[1],
                            "wiersze": row_numbers,
                        }
                        for outcome, row_numbers in variants.items()
                    ],
                }
            )
            continue
        outcome, row_numbers = next(iter(variants.items()))
        catalog[key] = {
            "powiat": outcome[0],
            "uwagi": outcome[1],
            "zrodlo": str(source_path),
            "rola_zrodla": source_role,
            "kolumna_klucza": key_column,
            "klucz_mapowania": key,
            "wiersze": row_numbers,
        }


def build_id_catalog(
    rows: list[dict[str, Any]],
    source_path: Path,
    conflicts: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    grouped: dict[str, dict[tuple[str, str], list[dict[str, Any]]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for row in rows:
        identifier = row.get("identyfikator", "")
        if identifier:
            grouped[identifier][(row["powiat"], row["uwagi"])].append(row)

    result: dict[str, dict[str, Any]] = {}
    for identifier, variants in grouped.items():
        if len(variants) != 1:
            conflicts.append(
                {
                    "rodzaj_konfliktu": "identyfikator",
                    "klucz": identifier,
                    "kolumna": "identyfikator",
                    "zrodlo": str(source_path),
                    "warianty": [
                        {
                            "powiat": outcome[0],
                            "uwagi": outcome[1],
                            "wiersze": [int(row["row_number"]) for row in rows_for_id],
                        }
                        for outcome, rows_for_id in variants.items()
                    ],
                }
            )
            continue
        outcome, rows_for_id = next(iter(variants.items()))
        accepted_keys = {
            normalize_lookup_key(row[column])
            for row in rows_for_id
            for column in ("nazwa_json", "nazwa_json_org")
            if row.get(column)
        }
        result[identifier] = {
            "powiat": outcome[0],
            "uwagi": outcome[1],
            "zrodlo": str(source_path),
            "rola_zrodla": "podstawowe_id",
            "kolumna_klucza": "identyfikator",
            "klucz_mapowania": identifier,
            "wiersze": [int(row["row_number"]) for row in rows_for_id],
            "akceptowane_nazwy": accepted_keys,
        }
    return result


def build_mapping_catalog(
    primary_path: Path,
    fallback_path: Path | None,
) -> dict[str, Any]:
    primary_rows, primary_metadata = read_mapping_rows(primary_path)
    conflicts: list[dict[str, Any]] = []
    catalog: dict[str, dict[str, Any]] = {}
    blocked_keys: set[str] = set()

    # Priorytet: glowna nazwa z nowego arkusza, alias OCR, a na koncu
    # brakujace klucze ze starszego arkusza.
    add_grouped_mappings(
        catalog, blocked_keys, conflicts,
        group_rows_by_key(primary_rows, "nazwa_json"),
        primary_path, "nazwa_json", "podstawowe",
    )
    add_grouped_mappings(
        catalog, blocked_keys, conflicts,
        group_rows_by_key(primary_rows, "nazwa_json_org"),
        primary_path, "nazwa_json_org", "alias_podstawowy",
    )
    id_catalog = build_id_catalog(primary_rows, primary_path, conflicts)

    metadata = [primary_metadata]
    if fallback_path is not None:
        fallback_rows, fallback_metadata = read_mapping_rows(fallback_path)
        metadata.append(fallback_metadata)
        add_grouped_mappings(
            catalog, blocked_keys, conflicts,
            group_rows_by_key(fallback_rows, "nazwa_json"),
            fallback_path, "nazwa_json", "rezerwowe",
        )
        add_grouped_mappings(
            catalog, blocked_keys, conflicts,
            group_rows_by_key(fallback_rows, "nazwa_json_org"),
            fallback_path, "nazwa_json_org", "alias_rezerwowy",
        )

    normalized_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for exact_key, correction in catalog.items():
        normalized_groups[normalize_lookup_key(exact_key)].append(correction)

    normalized_catalog: dict[str, dict[str, Any]] = {}
    normalized_conflicts: set[str] = set()
    for normalized_key, corrections in normalized_groups.items():
        outcomes = {(item["powiat"], item["uwagi"]) for item in corrections}
        if len(outcomes) == 1:
            normalized_catalog[normalized_key] = corrections[0]
        else:
            normalized_conflicts.add(normalized_key)
            conflicts.append(
                {
                    "rodzaj_konfliktu": "klucz_po_normalizacji",
                    "klucz": normalized_key,
                    "kolumna": "wiele",
                    "zrodlo": "wiele",
                    "warianty": [
                        {
                            "powiat": item["powiat"],
                            "uwagi": item["uwagi"],
                            "klucz_mapowania": item["klucz_mapowania"],
                            "zrodlo": item["zrodlo"],
                        }
                        for item in corrections
                    ],
                }
            )

    return {
        "exact": catalog,
        "exact_conflicts": blocked_keys,
        "normalized": normalized_catalog,
        "normalized_conflicts": normalized_conflicts,
        "by_id": id_catalog,
        "conflicts": conflicts,
        "metadata": metadata,
    }


def lookup_correction(
    catalog: dict[str, Any],
    record_id: str,
    source_value: str,
) -> tuple[dict[str, Any] | None, str]:
    normalized_source = normalize_lookup_key(source_value)
    by_id = catalog["by_id"].get(record_id)
    if by_id and normalized_source in by_id.get("akceptowane_nazwy", set()):
        return by_id, "identyfikator"
    if source_value in catalog["exact"]:
        return catalog["exact"][source_value], "dokladne"
    if source_value in catalog["exact_conflicts"]:
        return None, "konflikt_dokladnego_klucza"
    if normalized_source in catalog["normalized"]:
        return catalog["normalized"][normalized_source], "znormalizowane"
    if normalized_source in catalog["normalized_conflicts"]:
        return None, "konflikt_po_normalizacji"
    return None, "brak_mapowania"


def iter_targets(data: Any) -> Iterator[dict[str, Any]]:
    """Zwraca rekordy indywidualne i elementy zbiorcze wraz z metadanymi."""

    if not isinstance(data, list):
        raise ValueError("Glowny element JSON powinien byc lista rekordow")
    for record_index, record in enumerate(data):
        if not isinstance(record, dict):
            continue
        if record.get("rodzaj") == "indywidualne":
            yield {
                "target": record,
                "rodzaj_celu": "indywidualne",
                "parent_ID": None,
                "record_index": record_index,
                "element_index": None,
            }
        elif record.get("rodzaj") == "zbiorcze":
            elements = record.get("elementy", [])
            if not isinstance(elements, list):
                continue
            for element_index, element in enumerate(elements):
                if isinstance(element, dict):
                    yield {
                        "target": element,
                        "rodzaj_celu": "element",
                        "parent_ID": record.get("ID"),
                        "record_index": record_index,
                        "element_index": element_index,
                    }


def add_note(record: dict[str, Any], note: str) -> bool:
    if not note:
        return False
    current = str(record.get(NOTE_FIELD, "") or "").strip()
    if not current:
        record[NOTE_FIELD] = note
        return True
    parts = [part.strip() for part in current.split(";") if part.strip()]
    if note not in parts:
        record[NOTE_FIELD] = f"{current}; {note}"
        return True
    return False


def write_json_atomic(path: Path, data: Any) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8") as handle:
        json.dump(data, handle, ensure_ascii=False, indent=2)
        handle.write("\n")
    temporary.replace(path)


def write_csv_atomic(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)
    temporary.replace(path)


def format_duration(seconds: float) -> str:
    total = int(round(seconds))
    minutes, secs = divmod(total, 60)
    hours, minutes = divmod(minutes, 60)
    if hours:
        return f"{hours}h {minutes}m {secs}s"
    if minutes:
        return f"{minutes}m {secs}s"
    return f"{secs}s"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Uzupelnia brakujace powiat_ujednolicony i powiat_uwagi "
            "w jednym wskazanym pliku SGKP."
        )
    )
    parser.add_argument("input", type=Path, help="Zrodlowy plik JSON")
    parser.add_argument("-o", "--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument(
        "--primary-xlsx", type=Path, default=DEFAULT_PRIMARY_XLSX,
        help=f"Podstawowy arkusz mapowania. Domyslnie: {DEFAULT_PRIMARY_XLSX}",
    )
    parser.add_argument(
        "--fallback-xlsx", type=Path, default=DEFAULT_FALLBACK_XLSX,
        help=f"Rezerwowy arkusz mapowania. Domyslnie: {DEFAULT_FALLBACK_XLSX}",
    )
    parser.add_argument(
        "--no-fallback", action="store_true",
        help="Nie korzystaj z rezerwowego arkusza.",
    )
    parser.add_argument(
        "--scan-only", action="store_true",
        help="Przygotuj raporty, ale nie zapisuj poprawionego JSON-u.",
    )
    parser.add_argument(
        "--force", action="store_true",
        help="Pozwol zastapic istniejacy plik wynikowy.",
    )
    return parser.parse_args()


def main() -> int:
    started_at = time.monotonic()
    args = parse_args()
    fallback_path = None if args.no_fallback else args.fallback_xlsx

    for path, label in (
        (args.input, "pliku JSON"),
        (args.primary_xlsx, "podstawowego arkusza"),
        (fallback_path, "rezerwowego arkusza"),
    ):
        if path is not None and not path.is_file():
            print(f"Nie znaleziono {label}: {path}", file=sys.stderr)
            return 2

    args.output_dir.mkdir(parents=True, exist_ok=True)
    prefix = args.output_dir / f"{args.input.stem}_powiat_completion"
    output_json = args.output_dir / f"{args.input.stem}_powiaty_completed.json"
    changes_csv = prefix.with_suffix(".changes.csv")
    unmatched_csv = prefix.with_suffix(".unmatched.csv")
    conflicts_json = prefix.with_suffix(".conflicts.json")
    summary_json = prefix.with_suffix(".summary.json")
    if not args.scan_only and output_json.exists() and not args.force:
        print(
            f"Plik wynikowy juz istnieje: {output_json}. Uzyj --force, aby go zastapic.",
            file=sys.stderr,
        )
        return 2
    if output_json.resolve() == args.input.resolve():
        print("Plik wynikowy musi byc inny niz wejsciowy.", file=sys.stderr)
        return 2

    try:
        with args.input.open("r", encoding="utf-8") as handle:
            data = json.load(handle)
        catalog = build_mapping_catalog(args.primary_xlsx, fallback_path)
    except (OSError, json.JSONDecodeError, ValueError) as exc:
        print(f"Nie mozna przygotowac danych: {exc}", file=sys.stderr)
        return 2

    changes: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []
    candidates = matched = changed_records = 0
    filled_unified = added_notes = 0
    match_modes: dict[str, int] = defaultdict(int)

    try:
        for metadata in iter_targets(data):
            target = metadata["target"]
            source_value = clean_source_value(target.get(SOURCE_FIELD))
            if not source_value or has_value(target.get(TARGET_FIELD)):
                continue
            candidates += 1
            record_id = clean_source_value(target.get("ID"))
            correction, match_mode = lookup_correction(
                catalog, record_id, source_value
            )
            if correction is None:
                unmatched.append(
                    {
                        "ID": record_id,
                        "nazwa": target.get("nazwa"),
                        "rodzaj_celu": metadata["rodzaj_celu"],
                        "parent_ID": metadata["parent_ID"],
                        "powiat_ocr": source_value,
                        "powod": match_mode,
                    }
                )
                continue

            matched += 1
            match_modes[match_mode] += 1
            before_unified = target.get(TARGET_FIELD)
            before_note = target.get(NOTE_FIELD)
            changed = False
            if correction["powiat"]:
                target[TARGET_FIELD] = correction["powiat"]
                filled_unified += 1
                changed = True
            if add_note(target, correction["uwagi"]):
                added_notes += 1
                changed = True
            if changed:
                changed_records += 1
                changes.append(
                    {
                        "ID": record_id,
                        "nazwa": target.get("nazwa"),
                        "rodzaj_celu": metadata["rodzaj_celu"],
                        "parent_ID": metadata["parent_ID"],
                        "powiat_ocr": source_value,
                        "powiat_ujednolicony_przed": before_unified,
                        "powiat_ujednolicony_po": target.get(TARGET_FIELD),
                        "powiat_uwagi_przed": before_note,
                        "powiat_uwagi_po": target.get(NOTE_FIELD),
                        "tryb_dopasowania": match_mode,
                        "rola_zrodla": correction["rola_zrodla"],
                        "plik_mapowania": correction["zrodlo"],
                        "kolumna_klucza": correction["kolumna_klucza"],
                        "klucz_mapowania": correction["klucz_mapowania"],
                        "wiersze_arkusza": ",".join(
                            str(item) for item in correction["wiersze"]
                        ),
                    }
                )
    except ValueError as exc:
        print(f"Niepoprawna struktura JSON: {exc}", file=sys.stderr)
        return 2

    write_csv_atomic(
        changes_csv, changes,
        [
            "ID", "nazwa", "rodzaj_celu", "parent_ID", "powiat_ocr",
            "powiat_ujednolicony_przed", "powiat_ujednolicony_po",
            "powiat_uwagi_przed", "powiat_uwagi_po", "tryb_dopasowania",
            "rola_zrodla", "plik_mapowania", "kolumna_klucza",
            "klucz_mapowania", "wiersze_arkusza",
        ],
    )
    write_csv_atomic(
        unmatched_csv, unmatched,
        ["ID", "nazwa", "rodzaj_celu", "parent_ID", "powiat_ocr", "powod"],
    )
    write_json_atomic(
        conflicts_json,
        {
            "liczba_konfliktow": len(catalog["conflicts"]),
            "konflikty": catalog["conflicts"],
        },
    )
    if not args.scan_only:
        write_json_atomic(output_json, data)

    summary = {
        "plik_zrodlowy": str(args.input),
        "plik_wynikowy": None if args.scan_only else str(output_json),
        "tryb": "tylko_raport" if args.scan_only else "zapis_kopii",
        "arkusze": catalog["metadata"],
        "mapowania_dokladne": len(catalog["exact"]),
        "mapowania_po_normalizacji": len(catalog["normalized"]),
        "mapowania_po_id": len(catalog["by_id"]),
        "konflikty_mapowan": len(catalog["conflicts"]),
        "kandydaci": candidates,
        "dopasowane": matched,
        "dopasowane_bez_zmiany": matched - changed_records,
        "niedopasowane": len(unmatched),
        "zmienione_rekordy": changed_records,
        "uzupelnione_powiat_ujednolicony": filled_unified,
        "dodane_powiat_uwagi": added_notes,
        "tryby_dopasowania": dict(sorted(match_modes.items())),
        "czas_utc": datetime.now(timezone.utc).isoformat(),
        "czas_calkowity": format_duration(time.monotonic() - started_at),
        "pliki_raportow": {
            "zmiany_csv": str(changes_csv),
            "niedopasowane_csv": str(unmatched_csv),
            "konflikty_json": str(conflicts_json),
        },
    }
    write_json_atomic(summary_json, summary)

    if not args.scan_only:
        print(output_json)
    print(changes_csv)
    print(unmatched_csv)
    print(summary_json)
    print(
        f"Kandydaci={candidates}, dopasowane={matched}, "
        f"niedopasowane={len(unmatched)}, zmienione={changed_records}",
        file=sys.stderr,
    )
    print(f"Czas wykonania: {summary['czas_calkowity']}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
