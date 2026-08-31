#!/usr/bin/env python3
"""Weryfikuje i normalizuje powiat_ocr za pomoca Gemmy przez Ollama.

Skrypt analizuje rekordy indywidualne i elementy rekordow zbiorczych. Do
modelu trafiaja wylacznie cele z niepustym polem ``powiat_ocr``, ktore nie
maja klucza ``powiat_ujednolicony``. Plik zrodlowy nigdy nie jest
modyfikowany. Po korekcie ``powiat_ocr`` skrypt korzysta z arkuszy XLSX, aby
uzupelnic ``powiat_ujednolicony`` oraz ``powiat_uwagi``. Poprawki trafiaja do
nowej kopii JSON, a decyzje modelu i mapowania do plikow audytowych.
"""

from __future__ import annotations

import argparse
import ast
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
import copy
import csv
from datetime import datetime, timezone
import hashlib
import json
import os
from pathlib import Path
import re
import sys
import time
from typing import Any
import unicodedata
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

import openpyxl


DEFAULT_MODEL = "gemma4:31b-cloud"
DEFAULT_OLLAMA_URL = "http://localhost:11434"
DEFAULT_OUTPUT_DIR = Path("sgkp_weryfikacja/powiat_ocr")
DEFAULT_WORKERS = 1
DEFAULT_TIMEOUT = 300.0
DEFAULT_RETRIES = 5
DEFAULT_RETRY_DELAY = 30.0
PROMPT_VERSION = "powiat_ocr_verification_ollama_v2"
TRANSIENT_HTTP_CODES = {408, 425, 429, 500, 502, 503, 504}
CHANGE_STATUSES = {"zmiana_formy", "korekta_wartosci"}
ALL_STATUSES = {
    "bez_zmiany",
    "zmiana_formy",
    "korekta_wartosci",
    "brak_potwierdzenia",
    "niejednoznaczne",
}

BASE_PATH = Path(__file__).resolve().parent.parent
DEFAULT_PRIMARY_XLSX = (
    BASE_PATH
    / "dane_do_korekt"
    / "powiaty_ujednolicon 13.01.2026 r. wersja uzupełniona - nazwa_json_org.xlsx"
)
DEFAULT_FALLBACK_XLSX = BASE_PATH / "dane_do_korekt" / "powiaty_ujednolicone.xlsx"
UNIFIED_FIELD = "powiat_ujednolicony"
NOTE_FIELD = "powiat_uwagi"
NOTE_NOT_APPLICABLE = "nie dotyczy"

POWIAT_CUE_RE = re.compile(r"(?i)(?<!\w)(?:pow(?=\s|[.,:])|powiat\w*)")
POWIAT_PREFIX_RE = re.compile(r"(?i)^\s*(?:pow(?=\s|[.,:])|powiat\w*)")
GUBERNIA_PREFIX_RE = re.compile(
    r"(?i)^\s*(?:g[.]|gub(?=\s|[.,:])|gubern\w*)"
)
COUNTY_TOWN_RE = re.compile(
    r"(?ix)(?<!\w)"
    r"(?:m[.]?|mto|msto|miasto)\s+"
    r"(?:pow(?=\s|[.,:])|powiat\w*)[.,]?"
)
GUBERNIA_NAME_RE = re.compile(
    r"(?ix)(?<!\w)"
    r"(?:g[.]|gub[.]?|gubern\w*)\s+"
    r"[^\W\d_]+(?:[-–][^\W\d_]+)*"
)
SHARED_COUNTY_GOVERNORATE_RE = re.compile(
    r"(?ix)(?<!\w)"
    r"(?:pow(?=\s|[.,:])|powiat\w*)[.,]?\s*"
    r"(?:i|oraz)\s*"
    r"(?:g[.]|gub[.]?|gubern\w*)\s+"
    r"(?P<name>[^\W\d_]+(?:[-–][^\W\d_]+)*)"
)
POWIAT_NON_MEMBERSHIP_RE = re.compile(
    r"(?ix)"
    r"(?:\b(?:sąd\w*|urząd\w*|szkoł\w*|droga\w*|naczelnik\w*)\b"
    r".{0,60}\b(?:pow(?=\s|[.,:])|powiat\w*)\b)"
    r"|(?:\b(?:odległ\w*|wiorst\w*|mil\w*|kilometr\w*)\b"
    r".{0,80}\b(?:pow(?=\s|[.,:])|powiat\w*)\b)"
    r"|(?:\bpowiat\w*\b\s+(?:licz\w*|składa\w*|obejm\w*))"
)


class OllamaTransientError(RuntimeError):
    """Blad tymczasowy, po ktorym warto ponowic zapytanie."""


class ModelResponseError(ValueError):
    """Niepoprawna odpowiedz modelu wraz z surowa trescia."""

    def __init__(self, message: str, raw_output: str = "") -> None:
        super().__init__(message)
        self.raw_output = raw_output


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    with path.open("r", encoding="utf-8") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            if line.startswith("export "):
                line = line.removeprefix("export ").strip()
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip()
            if not key or key in os.environ:
                continue
            if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
                value = value[1:-1]
            os.environ[key] = value


def load_env_files(input_path: Path) -> None:
    candidates = [
        Path.cwd() / ".env",
        Path(__file__).resolve().parent / ".env",
        input_path.resolve().parent / ".env",
    ]
    seen: set[Path] = set()
    for path in candidates:
        if path not in seen:
            seen.add(path)
            load_env_file(path)


def load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def write_json(path: Path, data: Any) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8") as handle:
        json.dump(data, handle, ensure_ascii=False, indent=2)
        handle.write("\n")
    temporary.replace(path)


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not path.exists():
        return rows
    with path.open("r", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError as exc:
                raise RuntimeError(f"Niepoprawny JSONL {path}:{line_number}") from exc
            if isinstance(row, dict):
                rows.append(row)
    return rows


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False))
            handle.write("\n")
    temporary.replace(path)


def append_jsonl(path: Path, row: dict[str, Any]) -> None:
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(row, ensure_ascii=False))
        handle.write("\n")


def canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def has_value(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    return value.strip().casefold() not in {"", "/", "null", "none", "brak"}


def normalized_text(value: str) -> str:
    value = unicodedata.normalize("NFKC", value).casefold()
    value = value.replace("„", '"').replace("”", '"').replace("’", "'")
    return " ".join(value.split())


def word_tokens(value: str) -> list[str]:
    return re.findall(r"[^\W\d_]+", normalized_text(value), flags=re.UNICODE)


def words_roughly_match(left: str, right: str) -> bool:
    if left == right:
        return True
    shortest = min(len(left), len(right))
    if shortest < 4:
        return False
    common = 0
    for left_char, right_char in zip(left, right):
        if left_char != right_char:
            break
        common += 1
    return common >= max(4, shortest - 3)


def value_is_present_in_evidence(value: str, evidence: str) -> bool:
    value_words = word_tokens(value)
    evidence_words = word_tokens(evidence)
    if not value_words or not evidence_words:
        return False
    return all(
        any(words_roughly_match(word, candidate) for candidate in evidence_words)
        for word in value_words
    )


def values_roughly_match(left: str, right: str) -> bool:
    left_words = word_tokens(left)
    right_words = word_tokens(right)
    if len(left_words) != len(right_words) or not left_words:
        return False
    return all(
        words_roughly_match(left_word, right_word)
        for left_word, right_word in zip(left_words, right_words)
    )


def shared_county_governorate_supports(value: str, evidence: str) -> bool:
    """Obsluguje jawne konstrukcje typu `w pow. i gub. niżegorodzkiej`."""

    for match in SHARED_COUNTY_GOVERNORATE_RE.finditer(evidence):
        if value_is_present_in_evidence(value, match.group("name")):
            return True
    return False


def county_evidence_without_governorate_names(evidence: str) -> str:
    """Usuwa nazwy wystepujace wylacznie po oznaczeniu guberni."""

    return GUBERNIA_NAME_RE.sub(" ", evidence)


def is_same_name_county_town(
    task: dict[str, Any],
    proposed: str,
    evidence: str,
) -> bool:
    """Rozpoznaje SGKP-owe `Nazwa, m. pow. gub. X`.

    W strukturze projektu powiat_ocr dla takiego miasta przyjmuje nazwe hasla,
    o ile tekst nie podaje osobno innej, jawnej nazwy powiatu.
    """

    task_name = str(task.get("nazwa", "") or "").strip()
    return (
        bool(task_name)
        and normalized_text(proposed) == normalized_text(task_name)
        and bool(COUNTY_TOWN_RE.search(evidence))
    )


def format_duration(seconds: float) -> str:
    total_seconds = int(round(seconds))
    minutes, secs = divmod(total_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    if hours:
        return f"{hours}h {minutes}m {secs}s"
    if minutes:
        return f"{minutes}m {secs}s"
    return f"{secs}s"


def short_error(error: Exception) -> str:
    message = " ".join((str(error).strip() or error.__class__.__name__).split())
    return message if len(message) <= 500 else message[:497] + "..."


def task_label(task: dict[str, Any]) -> str:
    name = " ".join(str(task.get("nazwa", "") or "").split())
    return f"{task['ID']} {name}".strip()


def text_scoped_to_element(elements: list[Any], element_index: int) -> str:
    """Usuwa doklejona kopie kolejnego elementu tylko z kontekstu modelu."""

    current = elements[element_index]
    if not isinstance(current, dict):
        return ""
    text = str(current.get("text", "") or "").strip()
    cut_positions: list[int] = []
    for later in elements[element_index + 1 :]:
        if not isinstance(later, dict):
            continue
        later_text = str(later.get("text", "") or "").strip()
        if len(later_text) < 10:
            continue
        position = text.find(later_text)
        if position > 0:
            cut_positions.append(position)
    if cut_positions:
        return text[: min(cut_positions)].rstrip()
    return text


def collect_tasks(data: Any) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """Zwraca kandydatow i statystyke selekcji."""

    if not isinstance(data, list):
        raise RuntimeError("Plik wejsciowy powinien zawierac liste rekordow")
    tasks: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    stats = {
        "cele_lacznie": 0,
        "z_powiat_ocr": 0,
        "pominiete_z_powiat_ujednolicony": 0,
        "pominiete_bez_tekstu": 0,
        "kandydaci": 0,
    }

    def add_task(
        target: dict[str, Any],
        record_index: int,
        element_index: int | None,
        parent: dict[str, Any] | None,
        scoped_text: str | None = None,
    ) -> None:
        stats["cele_lacznie"] += 1
        current = target.get("powiat_ocr")
        if not has_value(current):
            return
        stats["z_powiat_ocr"] += 1
        # Warunek celowo dotyczy obecnosci klucza, a nie prawdziwosci wartosci.
        if "powiat_ujednolicony" in target:
            stats["pominiete_z_powiat_ujednolicony"] += 1
            return
        text = (
            scoped_text
            if scoped_text is not None
            else str(target.get("text", "") or "").strip()
        )
        if not text:
            stats["pominiete_bez_tekstu"] += 1
            return
        target_id = str(target.get("ID", "") or "").strip()
        if not target_id:
            raise RuntimeError(
                f"Brak ID dla rekordu {record_index}, elementu {element_index}"
            )
        if target_id in seen_ids:
            raise RuntimeError(f"Powtorzone ID kandydata: {target_id}")
        seen_ids.add(target_id)
        tasks.append(
            {
                "ID": target_id,
                "nazwa": target.get("nazwa"),
                "rodzaj_celu": "element" if element_index is not None else "indywidualne",
                "parent_ID": str(parent.get("ID", "") or "").strip() if parent else None,
                "record_index": record_index,
                "element_index": element_index,
                "powiat_ocr_przed": str(current).strip(),
                "text": text,
            }
        )
        stats["kandydaci"] += 1

    for record_index, record in enumerate(data):
        if not isinstance(record, dict):
            continue
        if record.get("rodzaj") == "indywidualne":
            add_task(record, record_index, None, None)
        elif record.get("rodzaj") == "zbiorcze":
            elements = record.get("elementy", [])
            if not isinstance(elements, list):
                continue
            for element_index, element in enumerate(elements):
                if isinstance(element, dict):
                    add_task(
                        element,
                        record_index,
                        element_index,
                        record,
                        text_scoped_to_element(elements, element_index),
                    )
    return tasks, stats


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "_", str(value or "").strip().casefold())


def normalize_lookup_key(value: Any) -> str:
    """Normalizuje klucz mapowania tylko w bezpiecznym zakresie."""

    return " ".join(str(value or "").split()).casefold()


def clean_mapping_value(value: Any) -> str:
    return " ".join(str(value or "").split()).strip()


def normalize_unified_county(value: Any) -> str:
    """Usuwa doprecyzowania techniczne, ale zachowuje nazwe powiatu."""

    text = str(value or "")
    text = re.sub(r"\s*\([^)]*\)", "", text)
    text = text.replace("(", "").replace(")", "")
    text = re.sub(
        r"\s*[-–]?\s*\bnie dotyczy\b", "", text, flags=re.IGNORECASE
    )
    return " ".join(text.split()).strip()


def normalize_mapping_note(value: Any) -> str:
    if isinstance(value, str) and re.search(
        r"\bnie dotyczy\b", value, flags=re.IGNORECASE
    ):
        return NOTE_NOT_APPLICABLE
    return ""


def parse_mapping_outcome(unified_value: Any, scope_value: Any) -> tuple[str, str]:
    """Zwraca nazwe ujednolicona i informacje o zakresie projektu."""

    county = normalize_unified_county(unified_value)
    note = normalize_mapping_note(scope_value) or normalize_mapping_note(unified_value)
    return county, note


def read_mapping_rows(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Wczytuje istotne kolumny arkusza i pomija jego sztuczny pusty ogon."""

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        first_row = next(
            sheet.iter_rows(min_row=1, max_row=1, max_col=50, values_only=True)
        )
        headers = {
            normalize_header(value): index
            for index, value in enumerate(first_row)
            if value is not None
        }
        required = {"nazwa_json", "nazwa_ujednolicona"}
        missing = sorted(required - set(headers))
        if missing:
            raise ValueError(
                f"Brak kolumn w arkuszu {path}: {', '.join(missing)}"
            )
        relevant_names = (
            "nazwa_json",
            "nazwa_json_org",
            "nazwa_ujednolicona",
            "dotyczy?",
            "identyfikator",
        )
        max_col = max(headers[name] for name in relevant_names if name in headers) + 1
        rows: list[dict[str, Any]] = []
        consecutive_empty = 0
        last_scanned_row = 1
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, max_col=max_col, values_only=True), start=2
        ):
            last_scanned_row = row_number

            def cell(column: str) -> Any:
                index = headers.get(column)
                return row[index] if index is not None and index < len(row) else None

            relevant_values = [cell(name) for name in relevant_names]
            if not any(
                value is not None and str(value).strip()
                for value in relevant_values
            ):
                consecutive_empty += 1
                # Oba dostarczone arkusze maja blednie zapisany wymiar ponad
                # miliona wierszy, mimo ze dane sa zwarte i koncza sie po
                # kilku tysiacach pozycji.
                if consecutive_empty >= 200:
                    break
                continue
            consecutive_empty = 0
            unified = cell("nazwa_ujednolicona")
            name_json = clean_mapping_value(cell("nazwa_json"))
            name_org = clean_mapping_value(cell("nazwa_json_org"))
            identifier = clean_mapping_value(cell("identyfikator"))
            if unified is None or not (name_json or name_org or identifier):
                continue
            county, note = parse_mapping_outcome(unified, cell("dotyczy?"))
            if not county and not note:
                continue
            rows.append(
                {
                    "row_number": row_number,
                    "nazwa_json": name_json,
                    "nazwa_json_org": name_org,
                    "identyfikator": identifier,
                    "powiat": county,
                    "uwagi": note,
                }
            )
        metadata = {
            "plik": str(path),
            "arkusz": sheet.title,
            "wiersze_mapowania": len(rows),
            "ostatni_sprawdzony_wiersz": last_scanned_row,
            "kolumny": sorted(headers),
        }
        return rows, metadata
    finally:
        workbook.close()


def group_mapping_rows(
    rows: list[dict[str, Any]],
    key_column: str,
) -> dict[str, dict[tuple[str, str], list[int]]]:
    groups: dict[str, dict[tuple[str, str], list[int]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for row in rows:
        key = row.get(key_column, "")
        if not key:
            continue
        groups[key][(row["powiat"], row["uwagi"])].append(int(row["row_number"]))
    return groups


def add_grouped_mappings(
    catalog: dict[str, dict[str, Any]],
    blocked_keys: set[str],
    conflicts: list[dict[str, Any]],
    groups: dict[str, dict[tuple[str, str], list[int]]],
    source_path: Path,
    key_column: str,
    source_role: str,
) -> None:
    """Dodaje tylko jednoznaczne mapowania zgodnie z kolejnoscia zrodel."""

    for key, variants in groups.items():
        if key in catalog or key in blocked_keys:
            continue
        if len(variants) != 1:
            blocked_keys.add(key)
            conflicts.append(
                {
                    "rodzaj_konfliktu": "dokladny_klucz",
                    "klucz": key,
                    "kolumna": key_column,
                    "zrodlo": str(source_path),
                    "warianty": [
                        {
                            "powiat": outcome[0],
                            "uwagi": outcome[1],
                            "wiersze": row_numbers,
                        }
                        for outcome, row_numbers in variants.items()
                    ],
                }
            )
            continue
        outcome, row_numbers = next(iter(variants.items()))
        catalog[key] = {
            "powiat": outcome[0],
            "uwagi": outcome[1],
            "zrodlo": str(source_path),
            "rola_zrodla": source_role,
            "kolumna_klucza": key_column,
            "klucz_mapowania": key,
            "wiersze": row_numbers,
        }


def build_id_mapping(
    rows: list[dict[str, Any]],
    source_path: Path,
    conflicts: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    grouped: dict[str, dict[tuple[str, str], list[dict[str, Any]]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for row in rows:
        identifier = row.get("identyfikator", "")
        if identifier:
            grouped[identifier][(row["powiat"], row["uwagi"])].append(row)

    result: dict[str, dict[str, Any]] = {}
    for identifier, variants in grouped.items():
        if len(variants) != 1:
            conflicts.append(
                {
                    "rodzaj_konfliktu": "identyfikator",
                    "klucz": identifier,
                    "kolumna": "identyfikator",
                    "zrodlo": str(source_path),
                    "warianty": [
                        {
                            "powiat": outcome[0],
                            "uwagi": outcome[1],
                            "wiersze": [int(row["row_number"]) for row in grouped_rows],
                        }
                        for outcome, grouped_rows in variants.items()
                    ],
                }
            )
            continue
        outcome, grouped_rows = next(iter(variants.items()))
        accepted_names = {
            normalize_lookup_key(row[column])
            for row in grouped_rows
            for column in ("nazwa_json", "nazwa_json_org")
            if row.get(column)
        }
        result[identifier] = {
            "powiat": outcome[0],
            "uwagi": outcome[1],
            "zrodlo": str(source_path),
            "rola_zrodla": "podstawowe_id",
            "kolumna_klucza": "identyfikator",
            "klucz_mapowania": identifier,
            "wiersze": [int(row["row_number"]) for row in grouped_rows],
            "akceptowane_nazwy": accepted_names,
        }
    return result


def build_mapping_catalog(
    primary_path: Path,
    fallback_path: Path | None,
) -> dict[str, Any]:
    primary_rows, primary_metadata = read_mapping_rows(primary_path)
    conflicts: list[dict[str, Any]] = []
    exact_catalog: dict[str, dict[str, Any]] = {}
    blocked_keys: set[str] = set()

    # Priorytet: nazwa mianownikowa z glownego arkusza, jej postac OCR,
    # nastepnie brakujace klucze z arkusza rezerwowego.
    add_grouped_mappings(
        exact_catalog,
        blocked_keys,
        conflicts,
        group_mapping_rows(primary_rows, "nazwa_json"),
        primary_path,
        "nazwa_json",
        "podstawowe",
    )
    add_grouped_mappings(
        exact_catalog,
        blocked_keys,
        conflicts,
        group_mapping_rows(primary_rows, "nazwa_json_org"),
        primary_path,
        "nazwa_json_org",
        "alias_podstawowy",
    )
    id_catalog = build_id_mapping(primary_rows, primary_path, conflicts)

    metadata = [primary_metadata]
    if fallback_path is not None:
        fallback_rows, fallback_metadata = read_mapping_rows(fallback_path)
        metadata.append(fallback_metadata)
        add_grouped_mappings(
            exact_catalog,
            blocked_keys,
            conflicts,
            group_mapping_rows(fallback_rows, "nazwa_json"),
            fallback_path,
            "nazwa_json",
            "rezerwowe",
        )
        add_grouped_mappings(
            exact_catalog,
            blocked_keys,
            conflicts,
            group_mapping_rows(fallback_rows, "nazwa_json_org"),
            fallback_path,
            "nazwa_json_org",
            "alias_rezerwowy",
        )

    normalized_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for exact_key, correction in exact_catalog.items():
        normalized_groups[normalize_lookup_key(exact_key)].append(correction)

    normalized_catalog: dict[str, dict[str, Any]] = {}
    normalized_conflicts: set[str] = set()
    for normalized_key, corrections in normalized_groups.items():
        outcomes = {(item["powiat"], item["uwagi"]) for item in corrections}
        if len(outcomes) == 1:
            normalized_catalog[normalized_key] = corrections[0]
        else:
            normalized_conflicts.add(normalized_key)
            conflicts.append(
                {
                    "rodzaj_konfliktu": "klucz_po_normalizacji",
                    "klucz": normalized_key,
                    "kolumna": "wiele",
                    "zrodlo": "wiele",
                    "warianty": [
                        {
                            "powiat": item["powiat"],
                            "uwagi": item["uwagi"],
                            "klucz_mapowania": item["klucz_mapowania"],
                            "zrodlo": item["zrodlo"],
                        }
                        for item in corrections
                    ],
                }
            )

    return {
        "exact": exact_catalog,
        "exact_conflicts": blocked_keys,
        "normalized": normalized_catalog,
        "normalized_conflicts": normalized_conflicts,
        "by_id": id_catalog,
        "conflicts": conflicts,
        "metadata": metadata,
    }


def lookup_mapping(
    catalog: dict[str, Any],
    record_id: str,
    source_value: str,
) -> tuple[dict[str, Any] | None, str]:
    normalized_source = normalize_lookup_key(source_value)
    by_id = catalog["by_id"].get(record_id)
    if by_id and normalized_source in by_id.get("akceptowane_nazwy", set()):
        return by_id, "identyfikator"
    if source_value in catalog["exact"]:
        return catalog["exact"][source_value], "dokladne"
    if source_value in catalog["exact_conflicts"]:
        return None, "konflikt_dokladnego_klucza"
    if normalized_source in catalog["normalized"]:
        return catalog["normalized"][normalized_source], "znormalizowane"
    if normalized_source in catalog["normalized_conflicts"]:
        return None, "konflikt_po_normalizacji"
    return None, "brak_mapowania"


def iter_targets(data: Any):
    if not isinstance(data, list):
        raise RuntimeError("Plik wejsciowy powinien zawierac liste rekordow")
    for record_index, record in enumerate(data):
        if not isinstance(record, dict):
            continue
        if record.get("rodzaj") == "indywidualne":
            yield record, {
                "rodzaj_celu": "indywidualne",
                "parent_ID": None,
                "record_index": record_index,
                "element_index": None,
            }
        elif record.get("rodzaj") == "zbiorcze":
            elements = record.get("elementy", [])
            if not isinstance(elements, list):
                continue
            for element_index, element in enumerate(elements):
                if isinstance(element, dict):
                    yield element, {
                        "rodzaj_celu": "element",
                        "parent_ID": record.get("ID"),
                        "record_index": record_index,
                        "element_index": element_index,
                    }


def add_mapping_note(target: dict[str, Any], note: str) -> bool:
    if not note:
        return False
    current = str(target.get(NOTE_FIELD, "") or "").strip()
    if not current:
        target[NOTE_FIELD] = note
        return True
    parts = [part.strip() for part in current.split(";") if part.strip()]
    if note not in parts:
        target[NOTE_FIELD] = f"{current}; {note}"
        return True
    return False


def apply_county_mapping(
    data: Any,
    catalog: dict[str, Any],
    allowed_ids: set[str] | None = None,
) -> dict[str, Any]:
    """Uzupelnia pola na podstawie powiat_ocr juz poprawionego przez model."""

    changes: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []
    candidates = matched = filled_unified = added_notes = 0
    match_modes: dict[str, int] = defaultdict(int)

    for target, metadata in iter_targets(data):
        source_value = clean_mapping_value(target.get("powiat_ocr"))
        if not source_value or has_value(target.get(UNIFIED_FIELD)):
            continue
        record_id = clean_mapping_value(target.get("ID"))
        if allowed_ids is not None and record_id not in allowed_ids:
            continue
        candidates += 1
        correction, match_mode = lookup_mapping(catalog, record_id, source_value)
        if correction is None:
            unmatched.append(
                {
                    "ID": record_id,
                    "nazwa": target.get("nazwa"),
                    "rodzaj_celu": metadata["rodzaj_celu"],
                    "parent_ID": metadata["parent_ID"],
                    "powiat_ocr": source_value,
                    "powod": match_mode,
                }
            )
            continue

        matched += 1
        match_modes[match_mode] += 1
        before_unified = target.get(UNIFIED_FIELD)
        before_note = target.get(NOTE_FIELD)
        changed = False
        if correction["powiat"]:
            target[UNIFIED_FIELD] = correction["powiat"]
            filled_unified += 1
            changed = True
        if add_mapping_note(target, correction["uwagi"]):
            added_notes += 1
            changed = True
        if changed:
            changes.append(
                {
                    "ID": record_id,
                    "nazwa": target.get("nazwa"),
                    "rodzaj_celu": metadata["rodzaj_celu"],
                    "parent_ID": metadata["parent_ID"],
                    "powiat_ocr": source_value,
                    "powiat_ujednolicony_przed": before_unified,
                    "powiat_ujednolicony_po": target.get(UNIFIED_FIELD),
                    "powiat_uwagi_przed": before_note,
                    "powiat_uwagi_po": target.get(NOTE_FIELD),
                    "tryb_dopasowania": match_mode,
                    "rola_zrodla": correction["rola_zrodla"],
                    "plik_mapowania": correction["zrodlo"],
                    "kolumna_klucza": correction["kolumna_klucza"],
                    "klucz_mapowania": correction["klucz_mapowania"],
                    "wiersze_arkusza": ",".join(
                        str(item) for item in correction["wiersze"]
                    ),
                }
            )

    return {
        "changes": changes,
        "unmatched": unmatched,
        "kandydaci": candidates,
        "dopasowane": matched,
        "niedopasowane": len(unmatched),
        "zmienione_rekordy": len(changes),
        "uzupelnione_powiat_ujednolicony": filled_unified,
        "dodane_powiat_uwagi": added_notes,
        "tryby_dopasowania": dict(sorted(match_modes.items())),
    }


def task_fingerprint(task: dict[str, Any]) -> str:
    payload = {
        "ID": task["ID"],
        "nazwa": task.get("nazwa"),
        "text": task["text"],
        "powiat_ocr_przed": task["powiat_ocr_przed"],
        "brak_powiat_ujednolicony": True,
    }
    return sha256_text(canonical_json(payload))


def validation_key(source_name: str, task: dict[str, Any], model: str) -> str:
    payload = {
        "source": source_name,
        "task_fingerprint": task_fingerprint(task),
        "model": model,
        "prompt_version": PROMPT_VERSION,
    }
    return sha256_text(canonical_json(payload))


def ollama_chat_url(base_url: str) -> str:
    normalized = base_url.rstrip("/")
    if normalized.endswith("/api/chat"):
        return normalized
    if normalized.endswith("/api"):
        return normalized + "/chat"
    return normalized + "/api/chat"


def ollama_chat(
    base_url: str,
    api_key: str | None,
    model: str,
    prompt: str,
    timeout: float,
) -> str:
    payload = {
        "model": model,
        "messages": [
            {
                "role": "system",
                "content": (
                    "Jestes asystentem historyka analizujacym tekst SGKP. "
                    "Oceniasz tylko informacje jawnie obecne w tekscie i "
                    "odpowiadasz wylacznie poprawnym JSON-em."
                ),
            },
            {"role": "user", "content": prompt},
        ],
        "format": "json",
        "stream": False,
        "options": {"temperature": 0},
    }
    headers = {"Content-Type": "application/json", "Accept": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    request = Request(
        ollama_chat_url(base_url),
        data=json.dumps(payload).encode("utf-8"),
        headers=headers,
        method="POST",
    )
    try:
        with urlopen(request, timeout=timeout) as response:
            response_text = response.read().decode("utf-8", errors="replace")
    except HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        message = f"Ollama zwrocila HTTP {exc.code}: {body}"
        if exc.code in TRANSIENT_HTTP_CODES:
            raise OllamaTransientError(message) from exc
        raise RuntimeError(message) from exc
    except (URLError, TimeoutError) as exc:
        raise OllamaTransientError(
            f"Nie mozna polaczyc sie z Ollama: {base_url}"
        ) from exc

    try:
        response_data = json.loads(response_text)
    except json.JSONDecodeError as exc:
        raise OllamaTransientError("Ollama zwrocila niepoprawny JSON HTTP") from exc
    if not isinstance(response_data, dict):
        raise OllamaTransientError("Odpowiedz Ollamy nie jest obiektem")
    if response_data.get("error"):
        raise RuntimeError(f"Ollama zwrocila blad: {response_data['error']}")
    message = response_data.get("message")
    content = message.get("content", "") if isinstance(message, dict) else ""
    if not isinstance(content, str) or not content.strip():
        raise OllamaTransientError("Ollama zwrocila pusta odpowiedz")
    return content.strip()


def parse_model_json(output: str) -> dict[str, Any]:
    stripped = output.strip()
    candidates = [stripped]
    start = stripped.find("{")
    end = stripped.rfind("}")
    if start != -1 and end >= start:
        candidates.append(stripped[start : end + 1])
    for candidate in list(candidates):
        if candidate.startswith("{{") and candidate.endswith("}}"):
            candidates.append(candidate[1:-1])
        if candidate.startswith("{{"):
            candidates.append(candidate[1:])
        if candidate.endswith("}}"):
            candidates.append(candidate[:-1])
    errors: list[str] = []
    for candidate in candidates:
        try:
            parsed = json.loads(candidate)
        except json.JSONDecodeError as exc:
            errors.append(str(exc))
            continue
        if isinstance(parsed, dict):
            return parsed
        errors.append("element glowny nie jest obiektem")
    for candidate in candidates:
        try:
            parsed = ast.literal_eval(candidate)
        except (SyntaxError, ValueError, TypeError) as exc:
            errors.append(str(exc))
            continue
        if isinstance(parsed, dict):
            return parsed
        errors.append("element glowny zapisu tolerancyjnego nie jest obiektem")
    detail = errors[-1] if errors else "brak obiektu"
    raise ModelResponseError(f"Nie mozna odczytac JSON modelu: {detail}", output)


def build_prompt(task: dict[str, Any]) -> str:
    return f"""Zweryfikuj pole `powiat_ocr` dla jednego hasla lub podhasla SGKP.

ID: {task['ID']}
Nazwa miejscowosci: {task.get('nazwa')}
Rodzaj: {task['rodzaj_celu']}
Obecne powiat_ocr: {json.dumps(task['powiat_ocr_przed'], ensure_ascii=False)}

TEKST:
{task['text']}

ZADANIE:
Ustal, czy obecne `powiat_ocr` jest nazwa powiatu glownej miejscowosci opisanej
w TEKSCIE i czy nazwa jest zapisana w mianowniku. Jezeli tekst podaje nazwe
w przypadku zaleznym, zwroc jej forme mianownikowa.

REGULY:
1. Korzystaj tylko z TEKSTU. Nie uzywaj wiedzy geograficznej ani historycznej
   spoza tekstu i nie zgaduj.
2. Powiat musi dotyczyc glownej miejscowosci tego hasla lub podhasla. Nie bierz
   nazwy powiatu sasiedniej miejscowosci ani innego numerowanego obiektu.
3. Za dowod uznaj bezposredni zapis przynaleznosci, np. `pow. rzeczyckiego`
   albo `w powiecie dryssieńskim`.
4. Konstrukcja `Nazwa, m. pow. gub. X` albo `Nazwa, miasto powiatowe guberni X`
   znaczy: Nazwa jest miastem powiatowym lezacym w GUBERNI X. Wyraz X jest
   nazwa guberni, NIE powiatu. Jezeli tekst nie podaje pozniej osobnej nazwy
   powiatu, w strukturze tych danych `powiat_ocr_po` ma byc rowne nazwie
   miejscowosci z pola `Nazwa miejscowosci`. Gdy obecne `powiat_ocr` jest juz
   rowne nazwie miejscowosci, zwroc `bez_zmiany`.
5. Nigdy nie tworz nazwy powiatu z przymiotnika stojacego po `g.`, `gub.` ani
   slowie `gubernia`. Dotyczy to szczegolnie zwrotow `pow. gub. kazańskiej`,
   `m. pow. gub. kałuskiej` itp. Nazwy `kazański` lub `kałuski` bylyby w tych
   przykladach bezpodstawnym przepisaniem nazwy guberni do pola powiatu.
6. Wyjatkiem jest jawny wspolny zapis `w pow. i gub. niżegorodzkiej`: wtedy
   przymiotnik odnosi sie zarowno do powiatu, jak i guberni. Takze osobne zdanie
   `Semenowski powiat leży...` jest bezposrednim dowodem nazwy powiatu i ma
   pierwszenstwo przed ogolnym zapisem o miescie powiatowym.
7. Samo wystapienie sadu/urzedu/szkoly powiatowej, odleglosci od miasta
   powiatowego lub statystyki calego powiatu nie okresla przynaleznosci.
8. `powiat_ocr_po` ma zawierac tylko nazwe powiatu, bez `pow.`, `powiat`, `gub.`
   i bez przecinka. Zachowaj historyczna pisownie oraz mozliwe bledy OCR;
   poprawiaj odmiane gramatyczna, a nie ujednolicenie nazwy do wspolczesnego
   slownika.
9. Dla form przymiotnikowych podaj mianownik liczby pojedynczej rodzaju
   meskiego: np. `pow. rzeczyckiego` -> `rzeczycki`, `w pow. dryssieńskim` ->
   `dryssieński`.
10. Nie zmieniaj kapitalizacji ani pisowni bez potrzeby. Jezeli obecna wartosc
   jest poprawna i juz w mianowniku, wybierz `bez_zmiany`.
11. Jezeli obecna wartosc wskazuje inny powiat, wybierz `korekta_wartosci` tylko
   wtedy, gdy TEKST jawnie podaje poprawna nazwe dla glownej miejscowosci.
12. Jezeli nie ma pewnego, bezposredniego dowodu, wybierz `brak_potwierdzenia`.
   Jezeli tekst daje sprzeczne lub rownorzedne mozliwosci, wybierz
   `niejednoznaczne`. W obu przypadkach nie proponuj zmiany.
13. `dowod` musi byc krotkim, doslownym cytatem z TEKSTU. Dla miasta
   powiatowego cytat powinien obejmowac nazwe miejscowosci i `m. pow.`. Dla
   innych przypadkow powinien obejmowac oznaczenie powiatu i jego nazwe.

PRZYKLADY KRYTYCZNE:
- `Czystopol, m. pow. gub. kazańskiej`, obecne `Czystopol` -> `bez_zmiany`,
  `powiat_ocr_po`: `Czystopol`. Nie wolno zwrocic `kazański`.
- `Kaszyn, miasto pow. gub. twerskiej`, obecne `gub. twerskiej` ->
  `korekta_wartosci`, `powiat_ocr_po`: `Kaszyn`. Nie wolno zwrocic `twerski`.
- `Kstowo, st. p. w pow. i gub. niżegorodzkiej` -> nazwa powiatu moze byc
  `niżegorodzki`, poniewaz slowo `i` jawnie laczy powiat i gubernie.
- `Semenowski powiat zajmuje północną część gubernii` -> nazwa powiatu to
  `Semenowski`, bo wystepuje bezposrednio przy slowie `powiat`.

Dozwolone statusy:
- `bez_zmiany`: wartosc poprawna i w mianowniku;
- `zmiana_formy`: ten sam powiat, lecz obecna nazwa wymaga zmiany na mianownik;
- `korekta_wartosci`: obecna nazwa jest bledna, a tekst jawnie podaje inna;
- `brak_potwierdzenia`: brak pewnego dowodu;
- `niejednoznaczne`: tekst nie pozwala wybrac jednej nazwy.

Zwroc wylacznie obiekt JSON:
{{
  "status": "zmiana_formy",
  "powiat_ocr_po": "rzeczycki",
  "dowod": "pow. rzeczyckiego",
  "uzasadnienie": "Ta sama nazwa jest w tekscie odmieniona przez przypadek."
}}

Dla `brak_potwierdzenia` i `niejednoznaczne` ustaw `powiat_ocr_po` na null.
Dla `bez_zmiany` powtorz obecna wartosc w `powiat_ocr_po`."""


def normalize_response(parsed: dict[str, Any], task: dict[str, Any]) -> dict[str, Any]:
    if isinstance(parsed.get("wynik"), dict):
        parsed = parsed["wynik"]
    status = str(parsed.get("status", "") or "").strip().casefold()
    aliases = {
        "zgodne": "bez_zmiany",
        "poprawne": "bez_zmiany",
        "bez zmian": "bez_zmiany",
        "odmieniona": "zmiana_formy",
        "zmiana odmiany": "zmiana_formy",
        "bledne": "korekta_wartosci",
        "błędne": "korekta_wartosci",
        "brak": "brak_potwierdzenia",
        "niepotwierdzone": "brak_potwierdzenia",
        "niejednoznaczny": "niejednoznaczne",
        "niejednoznaczna": "niejednoznaczne",
    }
    status = aliases.get(status, status)
    if status not in ALL_STATUSES:
        raise ValueError(f"Niepoprawny status: {status!r}")

    current = task["powiat_ocr_przed"]
    proposed_raw = parsed.get("powiat_ocr_po")
    proposed = str(proposed_raw or "").strip()
    evidence = str(parsed.get("dowod", "") or "").strip()
    reasoning = str(parsed.get("uzasadnienie", "") or "").strip()

    if status in {"brak_potwierdzenia", "niejednoznaczne"}:
        return {
            "status": status,
            "powiat_ocr_po": None,
            "dowod": evidence,
            "uzasadnienie": reasoning,
            "czy_zmienic": False,
        }

    if not evidence:
        raise ValueError("Brak dowodu dla decyzji o zweryfikowanej wartosci")
    if normalized_text(evidence) not in normalized_text(task["text"]):
        raise ValueError("Dowod nie wystepuje doslownie w tekscie")
    if not POWIAT_CUE_RE.search(evidence):
        raise ValueError("Dowod nie zawiera oznaczenia pow./powiat")
    if POWIAT_NON_MEMBERSHIP_RE.search(evidence):
        raise ValueError(
            "Dowod dotyczy instytucji, odleglosci lub statystyki powiatu, "
            "a nie jawnej przynaleznosci miejscowosci"
        )
    if not has_value(proposed):
        raise ValueError("Brak wartosci powiat_ocr_po")
    if POWIAT_PREFIX_RE.search(proposed):
        raise ValueError("powiat_ocr_po zawiera slowo lub skrot powiat")
    if GUBERNIA_PREFIX_RE.search(proposed):
        raise ValueError("powiat_ocr_po zawiera oznaczenie guberni")
    if "," in proposed:
        raise ValueError("powiat_ocr_po zawiera przecinek")

    same = normalized_text(proposed) == normalized_text(current)
    same_name_county_town = is_same_name_county_town(task, proposed, evidence)
    shared_county_governorate = shared_county_governorate_supports(
        proposed, evidence
    )
    county_evidence = county_evidence_without_governorate_names(evidence)
    direct_county_support = (
        bool(POWIAT_CUE_RE.search(county_evidence))
        and value_is_present_in_evidence(proposed, county_evidence)
    )
    if (
        COUNTY_TOWN_RE.search(evidence)
        and not same_name_county_town
        and len(POWIAT_CUE_RE.findall(county_evidence)) < 2
    ):
        # Sama nazwa miasta przed `m. pow.` nie jest dowodem utworzonej przez
        # model formy przymiotnikowej (np. Czehryn -> czehrynski). Drugi,
        # niezalezny zapis `Czehrynski powiat` nadal moze taka forme potwierdzic.
        direct_county_support = False
    if not (
        same_name_county_town
        or shared_county_governorate
        or direct_county_support
    ):
        raise ValueError(
            "Proponowana nazwa wystepuje tylko jako nazwa guberni albo nie "
            "ma bezposredniego potwierdzenia jako powiat"
        )
    evidence_for_county = evidence if shared_county_governorate else county_evidence

    if status == "bez_zmiany":
        if not same:
            raise ValueError("Status bez_zmiany, ale powiat_ocr_po rozni sie od obecnego")
        if not same_name_county_town and not value_is_present_in_evidence(
            current, evidence_for_county
        ):
            raise ValueError("Dowod nie zawiera obecnej nazwy powiatu")
        return {
            "status": status,
            "powiat_ocr_po": current,
            "dowod": evidence,
            "uzasadnienie": reasoning,
            "czy_zmienic": False,
        }

    if same:
        raise ValueError(f"Status {status}, ale wartosc nie ulegla zmianie")
    if not same_name_county_town and not value_is_present_in_evidence(
        proposed, evidence_for_county
    ):
        raise ValueError("Dowod nie zawiera proponowanej nazwy powiatu")
    if status == "zmiana_formy":
        if not values_roughly_match(current, proposed):
            raise ValueError(
                "Status zmiana_formy, ale obecna i proponowana nazwa nie "
                "wygladaja na formy tego samego wyrazu"
            )
        if not same_name_county_town and not value_is_present_in_evidence(
            current, evidence_for_county
        ):
            raise ValueError("Dowod nie zawiera obecnej, odmienionej nazwy powiatu")

    return {
        "status": status,
        "powiat_ocr_po": proposed,
        "dowod": evidence,
        "uzasadnienie": reasoning,
        "czy_zmienic": True,
    }


def extract_with_retries(
    base_url: str,
    api_key: str | None,
    model: str,
    task: dict[str, Any],
    timeout: float,
    retries: int,
    retry_delay: float,
) -> dict[str, Any]:
    prompt = build_prompt(task)
    current_prompt = prompt
    attempts = max(1, retries)
    for attempt in range(1, attempts + 1):
        output = ""
        try:
            output = ollama_chat(base_url, api_key, model, current_prompt, timeout)
            return normalize_response(parse_model_json(output), task)
        except OllamaTransientError as exc:
            if attempt == attempts:
                raise
            print(
                f"{task_label(task)}: blad tymczasowy ({attempt}/{attempts}): "
                f"{short_error(exc)}",
                file=sys.stderr,
            )
            time.sleep(max(0.0, retry_delay))
        except ValueError as exc:
            if attempt == attempts:
                raw_output = getattr(exc, "raw_output", output)
                raise ModelResponseError(short_error(exc), raw_output) from exc
            print(
                f"{task_label(task)}: bledny wynik modelu ({attempt}/{attempts}): "
                f"{short_error(exc)}",
                file=sys.stderr,
            )
            current_prompt = (
                prompt
                + "\n\nPOPRZEDNIA ODPOWIEDZ BYLA NIEPOPRAWNA. "
                + f"Walidator zglosil: {short_error(exc)}. "
                + "Zwroc caly poprawiony obiekt JSON bez komentarza.\n"
                + output
            )
            time.sleep(min(5.0, max(0.0, retry_delay)))
    raise RuntimeError("Nieudana odpowiedz modelu")


def make_result_row(
    source_path: Path,
    task: dict[str, Any],
    base_url: str,
    api_key: str | None,
    model: str,
    timeout: float,
    retries: int,
    retry_delay: float,
) -> dict[str, Any]:
    result = extract_with_retries(
        base_url,
        api_key,
        model,
        task,
        timeout,
        retries,
        retry_delay,
    )
    return {
        "plik_zrodlowy": source_path.name,
        "ID": task["ID"],
        "nazwa": task.get("nazwa"),
        "rodzaj_celu": task["rodzaj_celu"],
        "parent_ID": task.get("parent_ID"),
        "record_index": task["record_index"],
        "element_index": task.get("element_index"),
        "model": model,
        "prompt_version": PROMPT_VERSION,
        "validation_key": validation_key(source_path.name, task, model),
        "task_fingerprint": task_fingerprint(task),
        "powiat_ocr_przed": task["powiat_ocr_przed"],
        **result,
        "czas_utc": datetime.now(timezone.utc).isoformat(),
    }


def make_error_row(
    source_path: Path,
    task: dict[str, Any],
    model: str,
    error: Exception,
) -> dict[str, Any]:
    row: dict[str, Any] = {
        "plik_zrodlowy": source_path.name,
        "ID": task["ID"],
        "nazwa": task.get("nazwa"),
        "rodzaj_celu": task["rodzaj_celu"],
        "parent_ID": task.get("parent_ID"),
        "model": model,
        "prompt_version": PROMPT_VERSION,
        "validation_key": validation_key(source_path.name, task, model),
        "task_fingerprint": task_fingerprint(task),
        "powiat_ocr_przed": task["powiat_ocr_przed"],
        "blad": short_error(error),
        "czas_utc": datetime.now(timezone.utc).isoformat(),
    }
    raw_output = getattr(error, "raw_output", "")
    if raw_output:
        row["surowa_odpowiedz_modelu"] = raw_output
    return row


def write_candidates_csv(path: Path, tasks: list[dict[str, Any]]) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    columns = [
        "ID",
        "nazwa",
        "rodzaj_celu",
        "parent_ID",
        "powiat_ocr_przed",
    ]
    with temporary.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for task in tasks:
            writer.writerow({column: task.get(column) for column in columns})
    temporary.replace(path)


def write_results_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    columns = [
        "ID",
        "nazwa",
        "rodzaj_celu",
        "parent_ID",
        "powiat_ocr_przed",
        "status",
        "powiat_ocr_po",
        "czy_zmieniono",
        "dowod",
        "uzasadnienie",
    ]
    with temporary.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "ID": row.get("ID"),
                    "nazwa": row.get("nazwa"),
                    "rodzaj_celu": row.get("rodzaj_celu"),
                    "parent_ID": row.get("parent_ID"),
                    "powiat_ocr_przed": row.get("powiat_ocr_przed"),
                    "status": row.get("status"),
                    "powiat_ocr_po": row.get("powiat_ocr_po"),
                    "czy_zmieniono": row.get("czy_zmieniono", False),
                    "dowod": row.get("dowod"),
                    "uzasadnienie": row.get("uzasadnienie"),
                }
            )
    temporary.replace(path)


def write_mapping_changes_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    columns = [
        "ID",
        "nazwa",
        "rodzaj_celu",
        "parent_ID",
        "powiat_ocr",
        "powiat_ujednolicony_przed",
        "powiat_ujednolicony_po",
        "powiat_uwagi_przed",
        "powiat_uwagi_po",
        "tryb_dopasowania",
        "rola_zrodla",
        "plik_mapowania",
        "kolumna_klucza",
        "klucz_mapowania",
        "wiersze_arkusza",
        "czy_zastosowano",
    ]
    with temporary.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column) for column in columns})
    temporary.replace(path)


def write_mapping_unmatched_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    columns = ["ID", "nazwa", "rodzaj_celu", "parent_ID", "powiat_ocr", "powod"]
    with temporary.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column) for column in columns})
    temporary.replace(path)


def resolve_target(data: list[Any], row: dict[str, Any]) -> dict[str, Any]:
    record_index = int(row["record_index"])
    record = data[record_index]
    if not isinstance(record, dict):
        raise RuntimeError(f"Rekord {record_index} nie jest obiektem")
    element_index = row.get("element_index")
    if element_index is None:
        target = record
    else:
        elements = record.get("elementy")
        if not isinstance(elements, list):
            raise RuntimeError(f"Brak listy elementow dla {row['ID']}")
        target = elements[int(element_index)]
    if not isinstance(target, dict) or str(target.get("ID", "") or "") != row["ID"]:
        raise RuntimeError(f"Struktura pliku zmienila sie dla {row['ID']}")
    return target


def apply_results(data: Any, rows: list[dict[str, Any]]) -> int:
    if not isinstance(data, list):
        raise RuntimeError("Plik wejsciowy powinien zawierac liste rekordow")
    changed = 0
    for row in rows:
        row["czy_zmieniono"] = False
        if row.get("status") not in CHANGE_STATUSES or not row.get("czy_zmienic"):
            continue
        target = resolve_target(data, row)
        # Ponowne sprawdzenie chroni przed zastosowaniem starego cache do
        # zmienionego pliku oraz realizuje warunek uzytkownika przy samym zapisie.
        if "powiat_ujednolicony" in target:
            continue
        before = str(target.get("powiat_ocr", "") or "").strip()
        if before != str(row.get("powiat_ocr_przed", "") or "").strip():
            continue
        proposed = row.get("powiat_ocr_po")
        if not has_value(proposed) or normalized_text(str(proposed)) == normalized_text(before):
            continue
        target["powiat_ocr"] = str(proposed).strip()
        row["czy_zmieniono"] = True
        changed += 1
    return changed


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Weryfikuje i sprowadza powiat_ocr do mianownika przez Gemme/Ollama, "
            "ale tylko gdy rekord nie ma klucza powiat_ujednolicony."
        )
    )
    parser.add_argument("input", type=Path, help="Jeden zrodlowy plik sgkp_XX.json")
    parser.add_argument("-o", "--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--model", default=None)
    parser.add_argument("--ollama-url", default=None)
    parser.add_argument("--workers", type=int, default=DEFAULT_WORKERS)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument(
        "--ids",
        nargs="+",
        default=None,
        help="Przetworz tylko wskazane ID, rozdzielone spacjami lub przecinkami.",
    )
    parser.add_argument("--timeout", type=float, default=DEFAULT_TIMEOUT)
    parser.add_argument("--retries", type=int, default=DEFAULT_RETRIES)
    parser.add_argument("--retry-delay", type=float, default=DEFAULT_RETRY_DELAY)
    parser.add_argument(
        "--primary-xlsx",
        type=Path,
        default=DEFAULT_PRIMARY_XLSX,
        help=f"Podstawowy arkusz mapowania powiatow. Domyslnie: {DEFAULT_PRIMARY_XLSX}",
    )
    parser.add_argument(
        "--fallback-xlsx",
        type=Path,
        default=DEFAULT_FALLBACK_XLSX,
        help=f"Rezerwowy arkusz mapowania. Domyslnie: {DEFAULT_FALLBACK_XLSX}",
    )
    parser.add_argument(
        "--no-fallback",
        action="store_true",
        help="Nie korzystaj z rezerwowego arkusza powiatow.",
    )
    parser.add_argument(
        "--no-county-mapping",
        action="store_true",
        help="Nie uzupelniaj powiat_ujednolicony ani powiat_uwagi z XLSX.",
    )
    parser.add_argument(
        "--skip-failed",
        action="store_true",
        help="Nie ponawiaj zadan obecnych w aktualnym pliku errors.jsonl.",
    )
    parser.add_argument(
        "--scan-only",
        action="store_true",
        help="Tylko zapisz CSV kandydatow, bez wywolywania modelu.",
    )
    parser.add_argument(
        "--review-only",
        action="store_true",
        help="Uruchom model i raporty, ale nie stosuj zmian w kopii JSON.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Rozpocznij wyniki i bledy od nowa.",
    )
    args = parser.parse_args()
    if args.workers < 1:
        parser.error("--workers musi byc dodatnie")
    if args.limit is not None and args.limit < 1:
        parser.error("--limit musi byc dodatni")
    if args.retries < 1:
        parser.error("--retries musi byc dodatnie")
    if args.timeout <= 0:
        parser.error("--timeout musi byc dodatni")
    return args


def main() -> int:
    started_at = time.monotonic()
    args = parse_args()
    if not args.input.is_file():
        print(f"Nie znaleziono pliku: {args.input}", file=sys.stderr)
        return 2
    try:
        data = load_json(args.input)
        all_tasks, selection_stats = collect_tasks(data)
    except (OSError, json.JSONDecodeError, RuntimeError) as exc:
        print(f"Nie mozna przygotowac danych: {short_error(exc)}", file=sys.stderr)
        return 2

    selected_tasks = all_tasks
    if args.ids:
        requested_ids = {
            item.strip()
            for value in args.ids
            for item in value.split(",")
            if item.strip()
        }
        available_ids = {task["ID"] for task in all_tasks}
        missing_ids = sorted(requested_ids - available_ids)
        if missing_ids:
            print(
                "Wskazane ID nie sa kandydatami: " + ", ".join(missing_ids),
                file=sys.stderr,
            )
            return 2
        selected_tasks = [task for task in all_tasks if task["ID"] in requested_ids]

    args.output_dir.mkdir(parents=True, exist_ok=True)
    prefix = args.output_dir / f"{args.input.stem}_powiat_ocr_verification"
    candidates_csv = prefix.with_suffix(".candidates.csv")
    results_jsonl = prefix.with_suffix(".results.jsonl")
    results_csv = prefix.with_suffix(".results.csv")
    errors_jsonl = prefix.with_suffix(".errors.jsonl")
    summary_json = prefix.with_suffix(".summary.json")
    mapping_changes_csv = prefix.with_suffix(".mapping_changes.csv")
    mapping_unmatched_csv = prefix.with_suffix(".mapping_unmatched.csv")
    mapping_conflicts_json = prefix.with_suffix(".mapping_conflicts.json")
    output_json = args.output_dir / f"{args.input.stem}_powiat_ocr_normalized.json"
    write_candidates_csv(candidates_csv, selected_tasks)
    print(
        f"Kandydaci={len(all_tasks)}, wybrani={len(selected_tasks)}, "
        f"pominieci_z_powiat_ujednolicony="
        f"{selection_stats['pominiete_z_powiat_ujednolicony']}",
        file=sys.stderr,
    )
    if args.scan_only:
        print(candidates_csv)
        return 0

    mapping_catalog: dict[str, Any] | None = None
    if not args.no_county_mapping:
        fallback_path = None if args.no_fallback else args.fallback_xlsx
        for path, label in (
            (args.primary_xlsx, "podstawowego arkusza powiatow"),
            (fallback_path, "rezerwowego arkusza powiatow"),
        ):
            if path is not None and not path.is_file():
                print(f"Nie znaleziono {label}: {path}", file=sys.stderr)
                return 2
        try:
            mapping_catalog = build_mapping_catalog(args.primary_xlsx, fallback_path)
        except (OSError, ValueError) as exc:
            print(
                f"Nie mozna przygotowac mapowania powiatow: {short_error(exc)}",
                file=sys.stderr,
            )
            return 2

    load_env_files(args.input)
    model = args.model or os.environ.get("OLLAMA_MODEL", DEFAULT_MODEL)
    base_url = args.ollama_url or os.environ.get("OLLAMA_URL", DEFAULT_OLLAMA_URL)
    api_key = os.environ.get("OLLAMA_API_KEY")

    if args.overwrite:
        write_jsonl(results_jsonl, [])
        write_jsonl(errors_jsonl, [])
    previous_rows = load_jsonl(results_jsonl)
    existing = {
        str(row.get("validation_key", "") or ""): row
        for row in previous_rows
        if row.get("validation_key")
    }
    previous_errors = load_jsonl(errors_jsonl)
    existing_errors = {
        str(row.get("validation_key", "") or ""): row
        for row in previous_errors
        if row.get("validation_key")
    }
    for key in existing:
        existing_errors.pop(key, None)

    pending: list[dict[str, Any]] = []
    for task in selected_tasks:
        key = validation_key(args.input.name, task, model)
        if key in existing:
            continue
        if args.skip_failed and key in existing_errors:
            continue
        pending.append(task)
    if args.limit is not None:
        pending = pending[: args.limit]
    print(
        f"Zapisane wyniki={len(existing)}, do przetworzenia={len(pending)}",
        file=sys.stderr,
    )

    completed = 0
    new_errors = 0

    def handle_success(row: dict[str, Any]) -> None:
        nonlocal completed
        append_jsonl(results_jsonl, row)
        existing[row["validation_key"]] = row
        existing_errors.pop(row["validation_key"], None)
        completed += 1

    if args.workers == 1:
        for processed, task in enumerate(pending, start=1):
            try:
                row = make_result_row(
                    args.input,
                    task,
                    base_url,
                    api_key,
                    model,
                    args.timeout,
                    args.retries,
                    args.retry_delay,
                )
            except Exception as exc:
                new_errors += 1
                error_row = make_error_row(args.input, task, model, exc)
                existing_errors[error_row["validation_key"]] = error_row
                print(
                    f"[{processed}/{len(pending)}] Blad {task_label(task)}: "
                    f"{short_error(exc)}",
                    file=sys.stderr,
                )
            else:
                handle_success(row)
                print(
                    f"[{processed}/{len(pending)}] Gotowe {task_label(task)}: "
                    f"{row['status']} -> {row.get('powiat_ocr_po') or '-'}",
                    file=sys.stderr,
                )
    else:
        with ThreadPoolExecutor(max_workers=args.workers) as executor:
            futures = {
                executor.submit(
                    make_result_row,
                    args.input,
                    task,
                    base_url,
                    api_key,
                    model,
                    args.timeout,
                    args.retries,
                    args.retry_delay,
                ): task
                for task in pending
            }
            for processed, future in enumerate(as_completed(futures), start=1):
                task = futures[future]
                try:
                    row = future.result()
                except Exception as exc:
                    new_errors += 1
                    error_row = make_error_row(args.input, task, model, exc)
                    existing_errors[error_row["validation_key"]] = error_row
                    print(
                        f"[{processed}/{len(pending)}] Blad {task_label(task)}: "
                        f"{short_error(exc)}",
                        file=sys.stderr,
                    )
                else:
                    handle_success(row)
                    print(
                        f"[{processed}/{len(pending)}] Gotowe {task_label(task)}: "
                        f"{row['status']} -> {row.get('powiat_ocr_po') or '-'}",
                        file=sys.stderr,
                    )

    # Zachowujemy cache wszystkich nadal aktualnych kandydatow, nawet gdy
    # biezace uruchomienie zostalo ograniczone przez --ids lub --limit.
    current_keys = {
        validation_key(args.input.name, task, model) for task in all_tasks
    }
    final_rows = [existing[key] for key in current_keys if key in existing]
    final_rows.sort(key=lambda row: str(row.get("ID", "")))
    final_error_rows = [
        existing_errors[key]
        for key in current_keys
        if key in existing_errors and key not in existing
    ]
    final_error_rows.sort(key=lambda row: str(row.get("ID", "")))

    # Etap XLSX musi widziec powiat_ocr po korekcie modelu. W trybie
    # --review-only wykonujemy oba etapy na kopii roboczej, lecz zapisujemy
    # niezmieniona kopie danych zrodlowych oraz oznaczamy zmiany jako
    # niezastosowane.
    processed_data = copy.deepcopy(data)
    proposed_ocr_changes = apply_results(processed_data, final_rows)
    mapping_result: dict[str, Any] = {
        "changes": [],
        "unmatched": [],
        "kandydaci": 0,
        "dopasowane": 0,
        "niedopasowane": 0,
        "zmienione_rekordy": 0,
        "uzupelnione_powiat_ujednolicony": 0,
        "dodane_powiat_uwagi": 0,
        "tryby_dopasowania": {},
    }
    if mapping_catalog is not None:
        verified_ids = {str(row.get("ID", "") or "") for row in final_rows}
        mapping_result = apply_county_mapping(
            processed_data,
            mapping_catalog,
            allowed_ids=verified_ids,
        )

    for mapping_row in mapping_result["changes"]:
        mapping_row["czy_zastosowano"] = not args.review_only
    if args.review_only:
        changed = 0
        output_data = data
        for row in final_rows:
            row["czy_zmieniono"] = False
    else:
        changed = proposed_ocr_changes
        output_data = processed_data

    write_json(output_json, output_data)
    write_jsonl(results_jsonl, final_rows)
    write_jsonl(errors_jsonl, final_error_rows)
    write_results_csv(results_csv, final_rows)
    write_mapping_changes_csv(mapping_changes_csv, mapping_result["changes"])
    write_mapping_unmatched_csv(mapping_unmatched_csv, mapping_result["unmatched"])
    mapping_conflicts = mapping_catalog["conflicts"] if mapping_catalog else []
    write_json(
        mapping_conflicts_json,
        {
            "liczba_konfliktow": len(mapping_conflicts),
            "konflikty": mapping_conflicts,
        },
    )

    status_counts = {
        status: sum(row.get("status") == status for row in final_rows)
        for status in sorted(ALL_STATUSES)
    }
    summary = {
        "plik_zrodlowy": str(args.input),
        "plik_wynikowy": str(output_json),
        "model": model,
        "prompt_version": PROMPT_VERSION,
        "tryb": "tylko_raport" if args.review_only else "zastosowanie_do_kopii",
        "selekcja": selection_stats,
        "wybrani": len(selected_tasks),
        "nowe_zapytania": len(pending),
        "nowe_wyniki": completed,
        "nowe_bledy": new_errors,
        "wyniki_lacznie": len(final_rows),
        "bledy_lacznie": len(final_error_rows),
        "statusy": status_counts,
        "proponowane_zmiany_powiat_ocr": proposed_ocr_changes,
        "zmienione_powiat_ocr": changed,
        "mapowanie_powiatow": {
            "wlaczone": mapping_catalog is not None,
            "arkusze": mapping_catalog["metadata"] if mapping_catalog else [],
            "konflikty_mapowan": len(mapping_conflicts),
            "kandydaci": mapping_result["kandydaci"],
            "dopasowane": mapping_result["dopasowane"],
            "niedopasowane": mapping_result["niedopasowane"],
            "potencjalnie_uzupelnione_powiat_ujednolicony": mapping_result[
                "uzupelnione_powiat_ujednolicony"
            ],
            "potencjalnie_dodane_powiat_uwagi": mapping_result[
                "dodane_powiat_uwagi"
            ],
            "uzupelnione_powiat_ujednolicony": (
                0
                if args.review_only
                else mapping_result["uzupelnione_powiat_ujednolicony"]
            ),
            "dodane_powiat_uwagi": (
                0 if args.review_only else mapping_result["dodane_powiat_uwagi"]
            ),
            "tryby_dopasowania": mapping_result["tryby_dopasowania"],
        },
        "czas_calkowity": format_duration(time.monotonic() - started_at),
        "pliki": {
            "kandydaci_csv": str(candidates_csv),
            "wyniki_jsonl": str(results_jsonl),
            "wyniki_csv": str(results_csv),
            "bledy_jsonl": str(errors_jsonl),
            "mapowanie_zmiany_csv": str(mapping_changes_csv),
            "mapowanie_niedopasowane_csv": str(mapping_unmatched_csv),
            "mapowanie_konflikty_json": str(mapping_conflicts_json),
        },
    }
    write_json(summary_json, summary)
    print(output_json)
    print(results_csv)
    print(mapping_changes_csv)
    print(mapping_unmatched_csv)
    print(summary_json)
    print(
        f"Mapowanie XLSX: kandydaci={mapping_result['kandydaci']}, "
        f"dopasowane={mapping_result['dopasowane']}, "
        f"niedopasowane={mapping_result['niedopasowane']}, "
        f"powiat_ujednolicony={mapping_result['uzupelnione_powiat_ujednolicony']}, "
        f"powiat_uwagi={mapping_result['dodane_powiat_uwagi']}",
        file=sys.stderr,
    )
    print(f"Czas wykonania: {summary['czas_calkowity']}", file=sys.stderr)
    return 1 if new_errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
